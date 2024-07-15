from tkinter import *
import tkinter as tk
from tkinter.ttk import Combobox
from tkinter import messagebox
import openpyxl
from openpyxl import Workbook
import pathlib

root = Tk()
root.title("Data Entry")
root.geometry("700x400+300+200")
root.resizable(False, False)
root.configure(bg="#326273")

file_path = 'Backend_data.xlsx'
file = pathlib.Path(file_path)
if file.exists():
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
else:
    workbook = Workbook()
    sheet = workbook.active
    sheet['A1'] = "Full Name"
    sheet['B1'] = "Phone Number"
    sheet['C1'] = "Age"
    sheet['D1'] = "Gender"
    sheet['E1'] = "Address"
    workbook.save(file_path)


def submit():
    name = name_value.get()
    contact = contact_value.get()
    age = age_value.get()
    gender = gender_combobox.get()
    address = address_entry.get("1.0", END).strip()

    if not name or not contact or not age or not gender or not address:
        messagebox.showwarning("Warning", "All fields are required")
        return

    current_row = sheet.max_row
    sheet.cell(row=current_row + 1, column=1, value=name)
    sheet.cell(row=current_row + 1, column=2, value=contact)
    sheet.cell(row=current_row + 1, column=3, value=age)
    sheet.cell(row=current_row + 1, column=4, value=gender)
    sheet.cell(row=current_row + 1, column=5, value=address)

    workbook.save(file_path)
    messagebox.showinfo("Success", "Data saved successfully")
    clear()


def clear():
    name_value.set('')
    contact_value.set('')
    age_value.set('')
    gender_combobox.set('')
    address_entry.delete(1.0, END)


# Icon
icon_path = "C:/Python_Projects/Data_Entry_Form/logo.png"
if pathlib.Path(icon_path).exists():
    icon_image = PhotoImage(file=icon_path)
    root.iconphoto(False, icon_image)

# Heading
Label(root, text="Please fill out this Entry form", font="arial 13", bg="#326273", fg="#fff").place(x=20, y=20)

# Labels
Label(root, text='Name', font=23, bg="#326273", fg="#fff").place(x=50, y=100)
Label(root, text='Contact', font=23, bg="#326273", fg="#fff").place(x=50, y=150)
Label(root, text='Age', font=23, bg="#326273", fg="#fff").place(x=50, y=200)
Label(root, text='Gender', font=23, bg="#326273", fg="#fff").place(x=370, y=200)
Label(root, text='Address', font=23, bg="#326273", fg="#fff").place(x=50, y=250)

# Entries
name_value = StringVar()
contact_value = StringVar()
age_value = StringVar()
gender_value = StringVar()
address_value = StringVar()

name_entry = Entry(root, textvariable=name_value, width=45, bd=2, font=20)
contact_entry = Entry(root, textvariable=contact_value, width=45, bd=2, font=20)
age_entry = Entry(root, textvariable=age_value, width=15, bd=2, font=20)
address_entry = Text(root, width=62, height=4, bd=2)

name_entry.place(x=150, y=100)
contact_entry.place(x=150, y=150)
age_entry.place(x=150, y=200)
address_entry.place(x=150, y=250)

# Gender combobox
gender_combobox = Combobox(root, values=['Male', 'Female', 'Other'], font="arial 14", state="readonly", width=15)
gender_combobox.place(x=450, y=200)

# Buttons
Button(root, text="Submit", bg="#326273", fg="white", width=15, height=2, command=submit).place(x=200, y=350)
Button(root, text="Clear", bg="#326273", fg="white", width=15, height=2, command=clear).place(x=340, y=350)
Button(root, text="Exit", bg="#326273", fg="white", width=15, height=2, command=lambda: root.destroy()).place(x=480, y=350)

root.mainloop()
